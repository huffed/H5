from functools import lru_cache
from typing import Dict, Any, Optional

import openpyxl

wb: Optional[openpyxl.Workbook] = None
ws: Optional[Any] = None


@lru_cache(maxsize=None)
def load_workbook(filename: str):
    global wb, ws
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        return f"Error: File '{filename}' not found."
    except Exception as e:
        return f"An error occurred while loading the workbook: {str(e)}"


def get_data(start_row: int, end_row: int, start_column: int, end_column: str):
    end_column = end_column.upper()  # Convert input to uppercase for consistency
    number = 0
    for char in end_column:
        number = number * 26 + (ord(char) - ord('A')) + 1
    end_column = number

    if not wb:
        return "Error: Workbook has not been loaded. Please call 'load_workbook()' first."

    if not (start_row and end_row and start_column and end_column):
        return "Error: Missing parameters. 'start_row', 'end_row', 'start_column', and 'end_column' are required."

    try:
        data = {
            ws.cell(row, 1).value: {
                ws.cell(start_row, column).value: ws.cell(row, column).value
                for column in range(start_column, end_column + 1)
                if column > start_column or row > start_row
            }
            for row in range(start_row + 1, end_row + 1)
        }
    except Exception as e:
        return f"An error occurred while getting data: {str(e)}"
    return data


def get_schools(start_row: int = None, end_row: int = None, start_column: int = 2, end_column: str = None) -> \
        Dict[str, Dict[Any, Any]]:
    return get_data(start_row, end_row, start_column, end_column)


def get_instructors(start_row: int = None, end_row: int = None, start_column: int = 2, end_column: str = None) -> \
        Dict[str, Dict[Any, Any]]:
    return get_data(start_row, end_row, start_column, end_column)


def get_campsites(start_row: int = None, end_row: int = None, start_column: int = 2, end_column: str = None) -> \
        Dict[str, Dict[Any, Any]]:
    return get_data(start_row, end_row, start_column, end_column)


load_workbook("Mock database.xlsx")

__all__ = ["get_schools", "get_instructors", "get_campsites"]
