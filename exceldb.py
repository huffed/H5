from typing import Union, Dict, Any

import openpyxl


def get_data(worksheet, start_row: int, end_row: int,
             start_column: int, end_column: str) -> Union[Dict[str, Dict[str, Any]], str]:
    """
    Retrieves data from the given worksheet based on the specified range.

    :param worksheet: The worksheet object containing the data.
    :param start_row: The starting row index (1-based) for data retrieval.
    :param end_row: The ending row index (1-based) for data retrieval.
    :param start_column: The starting column index (1-based) for data retrieval.
    :param end_column: The ending column index (A-based or column letter) for data retrieval.

    :return: A dictionary containing the retrieved data with the specified structure, or an error message as a string.
    """
    worksheet = worksheet.sheet
    end_column = end_column.upper()
    number = 0
    for char in end_column:
        number = number * 26 + (ord(char) - ord('A')) + 1
    end_column = number

    if not worksheet:
        return "Error: Workbook has not been loaded. Please call 'load_workbook()' first."
    if not (start_row and end_row and start_column and end_column):
        return "Error: Missing parameters. 'start_row', 'end_row', 'start_column', and 'end_column' are required."

    try:
        data = {
            worksheet.cell(row, 1).value: {
                worksheet.cell(start_row, column).value: worksheet.cell(row, column).value
                for column in range(start_column, end_column + 1)
                if column > start_column or row > start_row
            }
            for row in range(start_row + 1, end_row + 1)
        }
    except Exception as e:
        raise Exception(f"An error occurred while getting data: {str(e)}")
    return data


class DataBase:
    def __init__(self, worksheet, start_row: int, end_row: int,
                 end_column: str, start_column: int = 2):
        """
        Represents a database created from the specified worksheet.

        :param worksheet: The worksheet object containing the data.
        :param start_row: The starting row index (1-based) for data retrieval.
        :param end_row: The ending row index (1-based) for data retrieval.
        :param end_column: The ending column index (A-based or column letter) for data retrieval.
        :param start_column: The starting column index (1-based) for data retrieval. (default: 2)
        """
        self.data = get_data(worksheet, start_row, end_row, start_column, end_column)
        self.index = list(self.data.keys())
        self.columns = list(self.data[self.index[0]].keys())


def load_workbook(filename: str):
    """
    Loads the workbook from the specified file and returns the active worksheet object.

    :param filename: The name of the workbook file to load.

    :return: The active worksheet object if the workbook is successfully loaded, or an error message as a string.
    """
    try:
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.active
        return worksheet
    except FileNotFoundError:
        raise FileNotFoundError(f"Error: File '{filename}' not found.")
    except Exception as e:
        raise Exception(f"An error occurred while loading the workbook: {str(e)}")


class Spreadsheet:
    def __init__(self, filename: str):
        """
        Represents a spreadsheet with the given filename.

        :param filename: The name of the spreadsheet file.
        """
        self.filename = filename
        self.sheet = load_workbook(self.filename)

    def __setitem__(self, name: str, database: DataBase):
        """
        Sets the specified database with the given name.

        :param name: The name of the database.
        :param database: The database object representing the database.

        :raises ValueError: If the database name contains invalid characters (e.g., space).
        """
        if " " in name:
            raise ValueError(f"Invalid character in database name ({name}) - space")
        setattr(self, name, database)

    def __getitem__(self, name: str) -> Union[Dict[str, Dict[str, Any]], None]:
        """
        Retrieves the specified database by name.

        :param name: The name of the database.

        :return: The database's data dictionary if it exists, None otherwise.

        :raises AttributeError: If the specified database does not exist.
        """
        if hasattr(self, name):
            return getattr(self, name).data
        else:
            raise AttributeError(f"'Spreadsheet' object has no attribute '{name}'")

    def __getattr__(self, attr: str) -> Union[Dict[str, Dict[str, Any]], None]:
        """
        Retrieves the specified database by name.

        :param attr: The name of the attribute.

        :return: The database's data dictionary if it exists, None otherwise.

        :raises AttributeError: If the specified attribute does not exist.
        """
        if hasattr(self, attr):
            return getattr(self, attr).data
        else:
            raise AttributeError(f"'Spreadsheet' object has no attribute '{attr}'")
