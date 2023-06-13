import openpyxl


# class GetExcelData:
#     self.wb = openpyxl.load_workbook("\Mock database.xlsx")
#     self.ws = self.wb.active()

#     def schoolDataBase(self):
#         Data = self.ws.cell(row=1, column=1)
#         return Data


# gt = GetExcelData()
# print(gt.schoolDataBase)

wb = openpyxl.load_workbook("Mock database.xlsx")
ws = wb.active

data = ws.cell(row=1, column=1)
print(data.value)
